"""Append missing device models from the supplied device-model workbook.

Existing device-model records are preserved.  The workbook is deduplicated by
canonical brand and normalized model name before new enabled records are added.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sqlite3
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_DB = ROOT_DIR / "mis_mvp" / "data" / "mis_mvp.sqlite3"

BRAND_ALIASES = {
    "苹果": "Apple",
    "华为": "Huawei",
    "oppo": "OPPO",
    "vivo": "vivo",
}


def canonical_brand(value: Any) -> str:
    brand = str(value or "").strip()
    return BRAND_ALIASES.get(brand.casefold(), brand)


def normalized(value: Any) -> str:
    return "".join(str(value or "").casefold().split())


def fingerprint(conn: sqlite3.Connection) -> str:
    rows = [dict(row) for row in conn.execute("SELECT * FROM device_models ORDER BY device_model_id")]
    body = json.dumps(rows, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(body.encode("utf-8")).hexdigest()


def read_workbook(path: Path) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = {str(value or "").strip(): index for index, value in enumerate(next(sheet.iter_rows(values_only=True)), start=0)}
        required = {"机型", "品牌"}
        if not required.issubset(headers):
            raise ValueError(f"表格缺少必要列：{', '.join(sorted(required - set(headers)))}")
        models: list[dict[str, str]] = []
        excluded: list[dict[str, Any]] = []
        for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            model = str(row[headers["机型"]] or "").strip()
            brand = canonical_brand(row[headers["品牌"]])
            creator = str(row[headers.get("添加人", -1)] or "").strip() if "添加人" in headers else ""
            if not model or not brand:
                excluded.append({"row": row_no, "model": model, "brand": brand, "reason": "缺少机型或品牌"})
                continue
            models.append({"row": str(row_no), "model": model, "brand": brand, "creator": creator})
        return models, excluded
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    workbook_path = args.workbook.resolve()
    db_path = args.db.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not db_path.is_file():
        raise SystemExit(f"Database not found: {db_path}")

    incoming, excluded = read_workbook(workbook_path)
    unique: dict[tuple[str, str], dict[str, str]] = {}
    for item in incoming:
        key = (normalized(item["brand"]), normalized(item["model"]))
        if key in unique:
            excluded.append({"row": int(item["row"]), "model": item["model"], "brand": item["brand"], "reason": "表格内重复"})
            continue
        unique[key] = item

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = db_path.with_name(f"{db_path.stem}.before-device-model-import-{timestamp}{db_path.suffix}")
    report_path = db_path.with_name(f"device-model-import-{timestamp}.json")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"Database integrity check failed before import: {integrity}")
        before_fingerprint = fingerprint(conn)
        existing_keys = {
            (normalized(row["brand"]), normalized(row["model_name"]))
            for row in conn.execute("SELECT brand, model_name FROM device_models")
        }
        additions = [item for key, item in unique.items() if key not in existing_keys]
        skipped_existing = [item for key, item in unique.items() if key in existing_keys]
        if not args.dry_run:
            shutil.copy2(db_path, backup_path)
            conn.execute("BEGIN IMMEDIATE")
            next_sort_order = int(conn.execute("SELECT COALESCE(MAX(sort_order), 0) FROM device_models").fetchone()[0]) + 1
            for offset, item in enumerate(additions):
                remark = f"导入自 {workbook_path.name}" + (f"；表格添加人：{item['creator']}" if item["creator"] else "")
                conn.execute(
                    """
                    INSERT INTO device_models
                    (brand, model_name, colors_json, capacities_json, model_numbers_json, enabled, sort_order, remark)
                    VALUES (?, ?, '[]', '[]', '[]', 1, ?, ?)
                    """,
                    (item["brand"], item["model"], next_sort_order + offset, remark),
                )
            duplicate_count = conn.execute(
                "SELECT COUNT(*) FROM (SELECT brand, model_name FROM device_models GROUP BY brand, model_name HAVING COUNT(*) > 1)"
            ).fetchone()[0]
            if duplicate_count:
                raise RuntimeError("Device-model uniqueness validation failed")
            integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
            if integrity != "ok":
                raise RuntimeError(f"Database integrity check failed after import: {integrity}")
            conn.commit()
        after_fingerprint = fingerprint(conn)
        report = {
            "workbook": str(workbook_path),
            "database": str(db_path),
            "dry_run": args.dry_run,
            "backup": None if args.dry_run else str(backup_path),
            "integrity_check": integrity,
            "device_model_fingerprint_before": before_fingerprint,
            "device_model_fingerprint_after": after_fingerprint,
            "source_rows": len(incoming),
            "unique_source_models": len(unique),
            "added": additions,
            "skipped_existing": skipped_existing,
            "excluded": excluded,
            "added_by_brand": dict(sorted(Counter(item["brand"] for item in additions).items())),
        }
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"backup": report["backup"], "report": str(report_path), "added": len(additions), "skipped_existing": len(skipped_existing), "excluded": len(excluded)}, ensure_ascii=False))
        return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
